"""
Extract line items from Excel (.xlsx / .xls) financial statements.

Handles:
  - Merged cells (expanded before reading)
  - Label column auto-detection
  - Year header row detection for CY / PY column identification
  - Both current-year and prior-year value columns
  - All sheets in the workbook
"""
import io
import re

import openpyxl


def _is_numeric_cell(value: str) -> bool:
    cleaned = re.sub(r'[₹$,\s()\-]', '', str(value))
    try:
        float(cleaned)
        return bool(cleaned)
    except ValueError:
        return False


def _is_short_generic(label: str) -> bool:
    s = label.strip()
    if len(s) <= 2:
        return True
    if re.fullmatch(r'[IVXivx]+\.?', s):
        return True
    if re.fullmatch(r'\([a-zA-Z]\)', s):
        return True
    return False


def _is_year_like(s: str) -> bool:
    return bool(
        re.fullmatch(r'20\d{2}[-–\/](?:20)?\d{2}', s.strip())
        or re.fullmatch(r'\d{1,2}[-\/]\d{1,2}[-\/]20\d{2}', s.strip())
        or re.fullmatch(r'20\d{2}', s.strip())
        or re.search(r'(?:march|mar)\s+20\d{2}', s.strip(), re.IGNORECASE)
    )


def _expand_merged_cells(ws) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        min_col = merged_range.min_col
        min_row = merged_range.min_row
        max_col = merged_range.max_col
        max_row = merged_range.max_row
        top_left = ws.cell(min_row, min_col).value
        # Must unmerge first — openpyxl won't allow setting value on a MergedCell
        ws.unmerge_cells(str(merged_range))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(r, c).value = top_left


def _detect_value_cols(
    rows: list[tuple], label_col: int
) -> tuple[int | None, int | None]:
    """
    Scan first 10 rows for a header row containing year-like strings.
    Returns (cy_col, py_col). Falls back to first/second numeric data cols.
    """
    for row in rows[:10]:
        cells = [str(c).strip() if c is not None else "" for c in row]
        year_cols = [
            ci for ci, c in enumerate(cells)
            if ci > label_col and _is_year_like(c)
        ]
        if len(year_cols) >= 2:
            return year_cols[0], year_cols[1]
        if len(year_cols) == 1:
            return year_cols[0], None

    # Fallback: detect from data rows
    for row in rows[1:10]:
        cells = [str(c).strip() if c is not None else "" for c in row]
        num_cols = [ci for ci, c in enumerate(cells) if ci > label_col and _is_numeric_cell(c)]
        if len(num_cols) >= 2:
            return num_cols[0], num_cols[1]

    return None, None


def extract_excel(file_bytes: bytes) -> list[dict]:
    """Extract label-value pairs from all sheets of an Excel workbook."""
    items: list[dict] = []
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    for ws in wb.worksheets:
        _expand_merged_cells(ws)
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Detect label column
        col_text_scores: dict[int, int] = {}
        for row in rows:
            for ci, cell in enumerate(row):
                s = str(cell).strip() if cell is not None else ""
                if s and not _is_numeric_cell(s) and len(s) > 3:
                    col_text_scores[ci] = col_text_scores.get(ci, 0) + 1

        if not col_text_scores:
            continue

        label_col = max(col_text_scores, key=col_text_scores.get)
        cy_col, py_col = _detect_value_cols(rows, label_col)

        for row in rows:
            cells = [str(c).strip() if c is not None else "" for c in row]
            if not any(cells):
                continue
            if label_col >= len(cells):
                continue

            label = cells[label_col]
            if not label or _is_numeric_cell(label) or _is_short_generic(label):
                continue

            # Current year value
            cy_val = ""
            if cy_col is not None and cy_col < len(cells) and _is_numeric_cell(cells[cy_col]):
                cy_val = cells[cy_col]
            if not cy_val:
                for ci in range(label_col + 1, min(label_col + 7, len(cells))):
                    if _is_numeric_cell(cells[ci]):
                        cy_val = cells[ci]
                        if cy_col is None:
                            cy_col = ci
                        break

            # Prior year value
            py_val = ""
            if py_col is not None and py_col < len(cells) and _is_numeric_cell(cells[py_col]):
                py_val = cells[py_col]
            if not py_val and cy_col is not None:
                for ci in range(cy_col + 1, min(cy_col + 4, len(cells))):
                    if ci < len(cells) and _is_numeric_cell(cells[ci]):
                        py_val = cells[ci]
                        break

            if cy_val:
                items.append({
                    "raw_label": label,
                    "raw_value": cy_val,
                    "page": None,
                    "bbox": None,
                    "ocr_confidence": None,
                })
            if py_val:
                items.append({
                    "raw_label": "previous year " + label,
                    "raw_value": py_val,
                    "page": None,
                    "bbox": None,
                    "ocr_confidence": None,
                })

    return items
