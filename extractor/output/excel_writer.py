"""
Write a structured, human-editable Excel file per client into finaloutput/.

This is the actual correction workflow: the dashboard only surfaces what's
wrong (read-only), the user opens this file directly in Excel to fix flagged
cells, and pipeline.revalidate_from_excel() re-reads it to refresh validation
without re-running extraction.
"""
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from extractor.mapping.schema import AOC4_FIELDS, SECTIONS

_SECTION_LABELS = SECTIONS

# Typical AOC-4 attachments — applicability varies by company, so this is a
# checklist for the user to confirm, not a hard validation rule.
_ATTACHMENT_CHECKLIST = [
    "Balance Sheet", "Profit & Loss Statement", "Cash Flow Statement",
    "Notes to Accounts", "Auditor's Report", "Board's / Directors' Report",
    "CSR Report (if applicable)", "AOC-2 — Related Party Transactions (if applicable)",
    "Consolidated Financial Statements (if applicable)",
]

_STATUS_FILL = {
    "MISSING": PatternFill("solid", fgColor="FFC7CE"),
    "LOW-CONFIDENCE": PatternFill("solid", fgColor="FFEB9C"),
    "TYPE-ERROR": PatternFill("solid", fgColor="FFC7CE"),
    "OK": PatternFill("solid", fgColor="C6EFCE"),
}

_FIELDS_HEADER = ["Section", "Field Key", "Label", "Value", "Confidence", "Source File", "Page", "Status", "Field Type"]

_TYPE_FILL = {
    "MANUAL": PatternFill("solid", fgColor="DDEBF7"),  # light blue — you fill these
    "MCA": PatternFill("solid", fgColor="EDEDED"),      # grey — portal fills these
}


def _safe_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", name).strip()


def _status_for(key: str, info: dict | None, validation: dict) -> str:
    if info is None or info.get("value") in (None, ""):
        return "MISSING" if key in validation.get("missing_mandatory", []) else "—"
    if any(k == key for k, _ in validation.get("type_errors", [])):
        return "TYPE-ERROR"
    if info.get("confidence") == "LOW":
        return "LOW-CONFIDENCE"
    return "OK"


def write_client_excel(result: dict, finaloutput_dir: Path) -> Path:
    finaloutput_dir.mkdir(parents=True, exist_ok=True)
    client = result.get("client_name", "Unknown")
    cin = result.get("cin") or "NO-CIN"
    out_path = finaloutput_dir / f"{_safe_filename(client)}_{cin}.xlsx"

    wb = Workbook()
    fields_ws = wb.active
    fields_ws.title = "Fields"
    _write_fields_sheet(fields_ws, result)
    _write_validation_sheet(wb.create_sheet("Validation"), result)
    _write_attachments_sheet(wb.create_sheet("Attachments"), result)
    _write_manual_sheet(wb.create_sheet("Manual Entry"), result)

    wb.save(out_path)
    return out_path


def _write_fields_sheet(ws, result: dict):
    ws.append(_FIELDS_HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    fields = result.get("fields", {})
    validation = result.get("validation", {})

    for fdef in AOC4_FIELDS:
        info = fields.get(fdef.key)
        status = _status_for(fdef.key, info, validation)
        ws.append([
            _SECTION_LABELS.get(fdef.section, fdef.section),
            fdef.key,
            fdef.label,
            info.get("value") if info else "",
            info.get("confidence") if info else "",
            info.get("source") if info else "",
            info.get("page") if info else "",
            status,
            fdef.source,
        ])
        row_i = ws.max_row
        fill = _STATUS_FILL.get(status)
        if fill:
            ws.cell(row=row_i, column=8).fill = fill
        type_fill = _TYPE_FILL.get(fdef.source)
        if type_fill:
            ws.cell(row=row_i, column=9).fill = type_fill

    for i, width in enumerate([34, 30, 40, 22, 12, 20, 6, 14, 11], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def _write_validation_sheet(ws, result: dict):
    v = result.get("validation", {})
    ws.append(["Check", "Result"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.append(["Missing mandatory fields", ", ".join(v.get("missing_mandatory", [])) or "None"])
    ws.append(["Low confidence fields", ", ".join(v.get("low_confidence", [])) or "None"])
    ws.append(["Type errors", "; ".join(f"{k}: {msg}" for k, msg in v.get("type_errors", [])) or "None"])
    ws.append(["Balance sheet check (Assets = Equity + Liabilities)", v.get("balance_check") or "Insufficient data"])
    ws.append(["P&L check (Revenue - Expenses ≈ PBT)", v.get("pnl_check") or "Insufficient data"])
    ws.append(["CIN", result.get("cin") or "Not found — check document text or rename folder"])
    ws.append(["CIN source", result.get("cin_source", "not_found")])
    for i, width in enumerate([45, 70], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_attachments_sheet(ws, result: dict):
    ws.append(["Found attachment files (from attachments/ folder)"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    attachments = result.get("attachments", [])
    if attachments:
        for att in attachments:
            ws.append([att["name"]])
    else:
        ws.append(["(none found — check the attachments/ folder)"])

    ws.append([])
    header_row = ws.max_row + 1
    ws.append(["Typical AOC-4 attachments (confirm applicability yourself)"])
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
    for item in _ATTACHMENT_CHECKLIST:
        ws.append([item])
    ws.column_dimensions["A"].width = 60


def _write_manual_sheet(ws, result: dict):
    """Focused checklist of every MANUAL field (not in any document — you fill
    these). Pre-fills any value already present so you only complete the blanks."""
    ws.append(["Section", "Field", "Field Key", "Value (fill in before filing)"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    fields = result.get("fields", {})
    for fdef in AOC4_FIELDS:
        if fdef.source != "MANUAL":
            continue
        existing = fields.get(fdef.key, {}).get("value") if fields.get(fdef.key) else ""
        ws.append([_SECTION_LABELS.get(fdef.section, fdef.section), fdef.label, fdef.key, existing or ""])
    for i, width in enumerate([34, 44, 30, 32], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def read_manual_values_from_excel(xlsx_path: Path) -> dict:
    """Read filled values from the 'Manual Entry' sheet — the fields the CS types
    once and we persist per client. Columns: Section, Field, Field Key, Value."""
    wb = load_workbook(xlsx_path, data_only=True)
    if "Manual Entry" not in wb.sheetnames:
        return {}
    ws = wb["Manual Entry"]
    out: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            continue
        key, value = row[2], row[3]
        if key and value not in (None, ""):
            out[str(key)] = value
    return out


def read_fields_from_excel(xlsx_path: Path) -> dict:
    """Read a (possibly user-edited) Fields sheet back into a fields dict for revalidation."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Fields"]
    fields: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        _, key, _, value = row[0], row[1], row[2], row[3] if len(row) > 3 else None
        confidence = row[4] if len(row) > 4 else None
        source = row[5] if len(row) > 5 else None
        page = row[6] if len(row) > 6 else None
        if not key or value in (None, ""):
            continue
        fields[str(key)] = {
            "value": value,
            "confidence": confidence or "MED",
            "source": source,
            "page": page,
        }
    return fields
